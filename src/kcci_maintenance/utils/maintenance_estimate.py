"""
Maintenance Estimate Generator

Generates O&M maintenance cost estimates by analyzing ticket distribution
across route legs/segments. Outputs an Excel workbook with:
- Tickets per leg/segment
- Ticket breakdowns by type, duration, work type
- Cost calculations (locate, maintenance, etc.)
- Summary statistics
"""

import logging
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional
from datetime import datetime
from zipfile import ZipFile
import xml.etree.ElementTree as ET

import geopandas as gpd
import pandas as pd
from shapely.geometry import Point, LineString
from shapely.ops import nearest_points

logger = logging.getLogger(__name__)


class MaintenanceEstimateGenerator:
    """Generates maintenance cost estimates by route leg."""

    # Ticket types that represent actual excavation work (should be counted)
    EXCAVATION_TICKET_TYPES = {
        'Normal',
        'Emergency',
        'DigUp',
    }

    # Ticket types that do NOT represent excavation (should be excluded)
    NON_EXCAVATION_TICKET_TYPES = {
        'Update',          # Re-marks only, no new excavation
        'No Response',     # No locator response
        'Cancellation',    # Cancelled tickets
        'Recall',          # Recalled tickets
        'Survey/Design',   # Planning/survey work only
        'Non-Compliant',   # Non-compliant tickets
    }

    def __init__(
        self,
        kmz_path: Path,
        buffer_distance_m: float = 500.0,
    ):
        """Initialize maintenance estimate generator.

        Args:
            kmz_path: Path to KMZ file with route legs
            buffer_distance_m: Buffer distance for assigning tickets to legs (default 500m)
        """
        self.kmz_path = kmz_path
        self.buffer_distance_m = buffer_distance_m

        # Load route legs
        self.route_legs = self._load_route_legs()

        logger.info(f"Loaded {len(self.route_legs)} route legs from {kmz_path}")

    def _load_route_legs(self) -> gpd.GeoDataFrame:
        """Load route legs from KMZ file.

        Returns:
            GeoDataFrame with route leg geometries and names
        """
        try:
            # Extract KML from KMZ
            with ZipFile(self.kmz_path, 'r') as zf:
                kml_files = [name for name in zf.namelist() if name.endswith('.kml')]
                if not kml_files:
                    raise ValueError(f"No KML file found in {self.kmz_path}")

                kml_content = zf.read(kml_files[0])

            # Parse KML to extract route legs
            root = ET.fromstring(kml_content)
            ns = {'kml': 'http://www.opengis.net/kml/2.2'}

            placemarks = root.findall('.//kml:Placemark', ns)

            legs = []
            for placemark in placemarks:
                name_elem = placemark.find('kml:name', ns)
                linestring_elem = placemark.find('.//kml:LineString', ns)

                if linestring_elem is not None:
                    coords_elem = linestring_elem.find('kml:coordinates', ns)
                    if coords_elem is not None:
                        # Parse coordinates (format: "lon,lat,alt lon,lat,alt ...")
                        coords_text = coords_elem.text.strip()
                        coords = []
                        for coord_str in coords_text.split():
                            parts = coord_str.split(',')
                            if len(parts) >= 2:
                                lon, lat = float(parts[0]), float(parts[1])
                                coords.append((lon, lat))

                        if coords:
                            name = name_elem.text if name_elem is not None else "Unnamed"
                            legs.append({
                                'name': name,
                                'geometry': LineString(coords)
                            })

            # Create GeoDataFrame
            gdf = gpd.GeoDataFrame(legs, crs='EPSG:4326')
            return gdf

        except Exception as e:
            logger.error(f"Failed to load route legs: {e}")
            raise

    def assign_tickets_to_legs(
        self,
        tickets_df: pd.DataFrame,
    ) -> pd.DataFrame:
        """Assign tickets to route legs based on proximity.

        Args:
            tickets_df: DataFrame with geocoded tickets
                Required columns: ticket_number, latitude, longitude

        Returns:
            DataFrame with added columns:
                - route_leg: Name of assigned route leg
                - distance_to_leg_m: Distance to nearest point on leg
        """
        # Convert tickets to GeoDataFrame
        tickets_gdf = gpd.GeoDataFrame(
            tickets_df,
            geometry=gpd.points_from_xy(tickets_df.longitude, tickets_df.latitude),
            crs='EPSG:4326'
        )

        # For each ticket, find nearest leg
        assignments = []

        for idx, ticket in tickets_gdf.iterrows():
            if ticket.geometry is None or pd.isna(ticket.latitude):
                assignments.append({
                    'route_leg': None,
                    'distance_to_leg_m': None
                })
                continue

            # Find nearest leg
            min_distance = float('inf')
            nearest_leg = None

            for leg_idx, leg in self.route_legs.iterrows():
                # Calculate distance (in degrees, then convert to meters)
                nearest_pt_on_leg, nearest_pt_from_ticket = nearest_points(
                    leg.geometry, ticket.geometry
                )

                distance_m = self._calculate_distance_meters(
                    ticket.latitude, ticket.longitude,
                    nearest_pt_on_leg.y, nearest_pt_on_leg.x
                )

                if distance_m < min_distance:
                    min_distance = distance_m
                    nearest_leg = leg['name']

            # Only assign if within buffer distance
            if min_distance <= self.buffer_distance_m:
                assignments.append({
                    'route_leg': nearest_leg,
                    'distance_to_leg_m': round(min_distance, 2)
                })
            else:
                assignments.append({
                    'route_leg': 'Unassigned',
                    'distance_to_leg_m': round(min_distance, 2)
                })

        # Add assignments to dataframe
        result_df = tickets_df.copy()
        result_df['route_leg'] = [a['route_leg'] for a in assignments]
        result_df['distance_to_leg_m'] = [a['distance_to_leg_m'] for a in assignments]

        return result_df

    def _calculate_distance_meters(
        self,
        lat1: float,
        lng1: float,
        lat2: float,
        lng2: float,
    ) -> float:
        """Calculate distance in meters between two lat/lng points using Haversine."""
        from math import radians, sin, cos, sqrt, atan2

        R = 6371000  # Earth radius in meters

        lat1_rad = radians(lat1)
        lat2_rad = radians(lat2)
        dlat = radians(lat2 - lat1)
        dlng = radians(lng2 - lng1)

        a = sin(dlat / 2) ** 2 + cos(lat1_rad) * cos(lat2_rad) * sin(dlng / 2) ** 2
        c = 2 * atan2(sqrt(a), sqrt(1 - a))

        return R * c

    def _filter_excavation_tickets(self, tickets_df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
        """Filter tickets to only include those representing actual excavation.

        Args:
            tickets_df: DataFrame with all tickets

        Returns:
            Tuple of (filtered_df, filter_stats)
        """
        original_count = len(tickets_df)

        # Check if ticket_type column exists
        if 'ticket_type' not in tickets_df.columns:
            logger.warning("No ticket_type column found - using all tickets")
            return tickets_df, {
                'total_tickets': original_count,
                'excavation_tickets': original_count,
                'excluded_tickets': 0,
                'excluded_by_type': {}
            }

        # Filter to excavation tickets only
        excavation_mask = tickets_df['ticket_type'].isin(self.EXCAVATION_TICKET_TYPES)
        excavation_tickets = tickets_df[excavation_mask].copy()
        excluded_tickets = tickets_df[~excavation_mask].copy()

        # Count excluded tickets by type
        excluded_by_type = {}
        if len(excluded_tickets) > 0:
            excluded_counts = excluded_tickets['ticket_type'].value_counts()
            excluded_by_type = excluded_counts.to_dict()

        filter_stats = {
            'total_tickets': original_count,
            'excavation_tickets': len(excavation_tickets),
            'excluded_tickets': len(excluded_tickets),
            'excluded_by_type': excluded_by_type
        }

        logger.info(f"Filtered tickets: {original_count:,} total → {len(excavation_tickets):,} excavation tickets")
        logger.info(f"Excluded {len(excluded_tickets):,} non-excavation tickets ({len(excluded_tickets)/original_count*100:.1f}%)")

        return excavation_tickets, filter_stats

    def _calculate_time_span_years(self, tickets_df: pd.DataFrame) -> tuple[float, dict]:
        """Calculate the time span of ticket data in years.

        Args:
            tickets_df: DataFrame with ticket data

        Returns:
            Tuple of (years_span, time_info_dict)
        """
        # Look for date columns
        date_col = None
        for col in ['Creation', 'creation', 'created_at', 'date', 'Date']:
            if col in tickets_df.columns:
                date_col = col
                break

        if date_col is None:
            logger.warning("No date column found - assuming 1 year of data")
            return 1.0, {'years': 1.0, 'method': 'assumed', 'note': 'No date column found'}

        try:
            # Parse dates (convert to UTC to handle mixed timezones)
            dates = pd.to_datetime(tickets_df[date_col], errors='coerce', utc=True)
            valid_dates = dates.dropna()

            if len(valid_dates) == 0:
                logger.warning("No valid dates found - assuming 1 year of data")
                return 1.0, {'years': 1.0, 'method': 'assumed', 'note': 'No valid dates'}

            min_date = valid_dates.min()
            max_date = valid_dates.max()

            # Calculate years span
            days_span = (max_date - min_date).days
            years_span = max(1.0, days_span / 365.25)  # Minimum 1 year

            # Get year range
            min_year = min_date.year
            max_year = max_date.year

            time_info = {
                'years': round(years_span, 2),
                'min_date': min_date,
                'max_date': max_date,
                'min_year': min_year,
                'max_year': max_year,
                'method': 'calculated',
                'note': f'{min_year}-{max_year} ({round(years_span, 1)} years)'
            }

            logger.info(f"Data time span: {years_span:.2f} years ({min_year}-{max_year})")
            return years_span, time_info

        except Exception as e:
            logger.warning(f"Error calculating time span: {e} - assuming 1 year")
            return 1.0, {'years': 1.0, 'method': 'error', 'note': str(e)}

    def _apply_arial_font(self, worksheet, max_row: int = None, max_col: int = None) -> None:
        """Apply Arial font to all cells in a worksheet.

        Args:
            worksheet: Worksheet to apply font to
            max_row: Maximum row to apply font to (default: worksheet max_row)
            max_col: Maximum column to apply font to (default: worksheet max_column)
        """
        from openpyxl.styles import Font

        if max_row is None:
            max_row = worksheet.max_row
        if max_col is None:
            max_col = worksheet.max_column

        arial_font = Font(name='Arial', size=11)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                # Preserve bold and size attributes if they exist
                if cell.font and (cell.font.bold or cell.font.size):
                    cell.font = Font(
                        name='Arial',
                        size=cell.font.size if cell.font.size else 11,
                        bold=cell.font.bold if cell.font.bold else False
                    )
                else:
                    cell.font = arial_font

    def _write_quote_sheet(
        self,
        writer: pd.ExcelWriter,
        leg_details_df: pd.DataFrame,
        leg_row_mapping: dict,
        total_cost_row: int,
        total_cost_lease_row: int,
        project_name: str = "Project",
    ) -> None:
        """Write Quote sheet with pricing summary.

        Args:
            writer: Excel writer
            leg_details_df: DataFrame with leg details
            leg_row_mapping: Dict mapping leg names to row numbers in Leg Details sheet
            total_cost_row: Row number of Total Annual O&M Cost in Maintenance Estimate sheet
            total_cost_lease_row: Row number of Total Cost of Lease in Maintenance Estimate sheet
            project_name: Project name for table name
        """
        from openpyxl.styles import Font

        worksheet = writer.book.create_sheet('Quote', 0)  # Insert as first sheet

        # Set column widths to match reference
        worksheet.column_dimensions['A'].width = 26.71
        worksheet.column_dimensions['B'].width = 20.43
        worksheet.column_dimensions['C'].width = 28.14
        worksheet.column_dimensions['D'].width = 43.0
        worksheet.column_dimensions['E'].width = 34.43
        worksheet.column_dimensions['F'].width = 14.43
        worksheet.column_dimensions['G'].width = 8.86

        # Headers
        worksheet['A1'] = 'Leg'
        worksheet['B1'] = 'Length (miles)'
        worksheet['C1'] = 'Est. Distance (km)'
        worksheet['D1'] = 'Non-Recurring Charges (O&M NRC)'
        worksheet['E1'] = 'Monthly O&M Fee (MRC)'
        worksheet['F1'] = 'Leg % of Route'

        # Set row heights to match reference
        for row in range(1, 7):
            worksheet.row_dimensions[row].height = 22.5

        # Data rows - one per route leg
        row = 2
        data_rows = []
        for idx, leg_row in leg_details_df.iterrows():
            leg_name = leg_row['Route Leg']
            if leg_name == 'Unassigned':
                continue

            data_rows.append(row)
            worksheet[f'A{row}'] = leg_name

            # Length in miles - reference from Leg Details sheet
            worksheet[f'B{row}'] = f"='Leg Details'!B{row}"

            # Convert to km
            worksheet[f'C{row}'] = f'=B{row}*1.60934'

            # NRC - calculated as Inputs!B12 * Quote!F{row}
            worksheet[f'D{row}'] = f'=Inputs!$B$12*Quote!$F{row}'
            worksheet[f'D{row}'].number_format = '"$"#,##0.00'

            # MRC - calculated as (Maintenance Estimate Total / 12) * Leg % of Route
            worksheet[f'E{row}'] = f"='Maintenance Estimate'!$D${total_cost_row}/12*Quote!$F{row}"
            worksheet[f'E{row}'].number_format = '"$"#,##0.00'

            # Leg % of Route
            worksheet[f'F{row}'] = f'=Quote!$B{row}/$B$6'
            worksheet[f'F{row}'].number_format = '0.00%'

            row += 1

        # Total row
        total_row = 6  # Fixed row for total
        worksheet[f'A{total_row}'] = 'Total'
        worksheet[f'A{total_row}'].font = Font(name="Arial", bold=True)

        worksheet[f'B{total_row}'] = f'=SUM(B2:B{total_row-1})'
        worksheet[f'B{total_row}'].font = Font(name="Arial", bold=True)

        worksheet[f'C{total_row}'] = f'=SUM(C2:C{total_row-1})'
        worksheet[f'C{total_row}'].font = Font(name="Arial", bold=True)

        worksheet[f'D{total_row}'] = f'=Inputs!$B$12*Quote!$F{total_row}'
        worksheet[f'D{total_row}'].number_format = '"$"#,##0.00'
        worksheet[f'D{total_row}'].font = Font(name="Arial", bold=True)

        worksheet[f'E{total_row}'] = f"='Maintenance Estimate'!$D${total_cost_row}/12*Quote!$F{total_row}"
        worksheet[f'E{total_row}'].number_format = '"$"#,##0.00'
        worksheet[f'E{total_row}'].font = Font(name="Arial", bold=True)

        worksheet[f'F{total_row}'] = f'=Quote!$B{total_row}/$B${total_row}'
        worksheet[f'F{total_row}'].number_format = '0.00%'
        worksheet[f'F{total_row}'].font = Font(name="Arial", bold=True)

        # Summary calculations below the table
        summary_row = 8

        worksheet[f'D{summary_row}'] = 'Cost Per Foot/Month'
        worksheet[f'E{summary_row}'] = f'=E6/B6'
        worksheet[f'E{summary_row}'].number_format = '"$"#,##0.00'
        worksheet[f'E{summary_row}'].font = Font(name="Arial", bold=True)

        worksheet[f'D{summary_row+1}'] = 'Cost Per Kilometer/Month'
        worksheet[f'E{summary_row+1}'] = f'=E6/C6'
        worksheet[f'E{summary_row+1}'].number_format = '"$"#,##0.00'

        worksheet[f'D{summary_row+2}'] = 'Annual Cost'
        worksheet[f'E{summary_row+2}'] = f"='Maintenance Estimate'!D{total_cost_row}"
        worksheet[f'E{summary_row+2}'].number_format = '"$"#,##0.00'
        worksheet[f'E{summary_row+2}'].font = Font(name="Arial", bold=True)

        worksheet[f'D{summary_row+3}'] = 'Total Cost Right of Use 20 Year Lease'
        worksheet[f'E{summary_row+3}'] = f'=Quote!$D$6+\'Maintenance Estimate\'!D{total_cost_lease_row}'
        worksheet[f'E{summary_row+3}'].number_format = '"$"#,##0.00'
        worksheet[f'E{summary_row+3}'].font = Font(name="Arial", bold=True)

        worksheet[f'D{summary_row+4}'] = 'TCO Percentage of Installation Cost'
        worksheet[f'E{summary_row+4}'] = f'=E{summary_row+3}/Inputs!B10'
        worksheet[f'E{summary_row+4}'].number_format = '0.00%'
        worksheet[f'E{summary_row+4}'].font = Font(name='Arial', bold=True)

        # Apply Arial font to all cells
        self._apply_arial_font(worksheet, max_row=summary_row+4, max_col=7)

    def generate_estimate(
        self,
        tickets_df: pd.DataFrame,
        output_path: Path,
        project_name: str = "Project",
    ) -> None:
        """Generate maintenance estimate Excel workbook.

        Args:
            tickets_df: DataFrame with geocoded tickets
            output_path: Path for output Excel file
            project_name: Name of project for report title
        """
        logger.info(f"Generating maintenance estimate for {len(tickets_df)} tickets")

        # Calculate time span of data for annualization
        years_span, time_info = self._calculate_time_span_years(tickets_df)

        # Filter to excavation tickets only
        excavation_tickets, filter_stats = self._filter_excavation_tickets(tickets_df)

        # Assign tickets to legs
        tickets_with_legs = self.assign_tickets_to_legs(excavation_tickets)

        # Add annualization info to filter stats
        filter_stats['years_span'] = years_span
        filter_stats['time_info'] = time_info

        # Generate statistics (with annualization)
        summary_df = self._generate_summary_stats(tickets_with_legs, filter_stats)
        leg_details_df = self._generate_leg_details(tickets_with_legs, years_span)
        breakdown_df = self._generate_breakdowns(tickets_with_legs)
        cost_projections_df = self._generate_cost_projections(tickets_with_legs, leg_details_df)

        # Create Excel workbook
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # First write Maintenance Estimate to get the row mappings
            sheet_refs = self._write_summary_sheet(writer, summary_df, leg_details_df, project_name)

            # Sheet 2: Inputs (centralized parameters with annualization)
            self._write_inputs_sheet(writer, leg_details_df, filter_stats)

            # Sheet 3: Leg Details
            leg_details_df.to_excel(writer, sheet_name='Leg Details', index=False)

            # Apply percentage formatting to Leg Details sheet
            ws_leg_details = writer.book['Leg Details']
            for row in range(2, len(leg_details_df) + 2):  # Start at row 2 (after header)
                if 'Emergency %' in leg_details_df.columns:
                    col_idx = list(leg_details_df.columns).index('Emergency %') + 1
                    ws_leg_details.cell(row=row, column=col_idx).number_format = '0.0%'
                if 'Avg Confidence' in leg_details_df.columns:
                    col_idx = list(leg_details_df.columns).index('Avg Confidence') + 1
                    ws_leg_details.cell(row=row, column=col_idx).number_format = '0.00%'

            # Apply Arial font to Leg Details
            self._apply_arial_font(ws_leg_details)

            # Sheet 4: Cost Projections (with formulas referencing Inputs)
            self._write_cost_projections_sheet(writer, cost_projections_df)

            # Now write Quote sheet (after Leg Details is created since it references it)
            self._write_quote_sheet(
                writer,
                leg_details_df,
                sheet_refs['leg_row_mapping'],
                sheet_refs['total_cost_row'],
                sheet_refs['total_cost_lease_row'],
                project_name
            )

            # Sheet 5: Breakdowns
            breakdown_df.to_excel(writer, sheet_name='Ticket Breakdowns', index=False)

            # Apply percentage formatting to Ticket Breakdowns sheet
            ws_breakdowns = writer.book['Ticket Breakdowns']
            if 'Percentage' in breakdown_df.columns:
                col_idx = list(breakdown_df.columns).index('Percentage') + 1
                for row in range(2, len(breakdown_df) + 2):  # Start at row 2 (after header)
                    ws_breakdowns.cell(row=row, column=col_idx).number_format = '0.0%'

            # Apply Arial font to Ticket Breakdowns
            self._apply_arial_font(ws_breakdowns)

            # Sheet 6: Raw Data (tickets with leg assignments)
            tickets_with_legs.to_excel(writer, sheet_name='Ticket Assignments', index=False)

            # Apply Arial font to Ticket Assignments
            ws_assignments = writer.book['Ticket Assignments']
            self._apply_arial_font(ws_assignments)

        logger.info(f"Maintenance estimate saved to {output_path}")

    def _generate_summary_stats(self, tickets_df: pd.DataFrame, filter_stats: dict = None) -> pd.DataFrame:
        """Generate overall summary statistics.

        Args:
            tickets_df: DataFrame with filtered tickets (excavation only)
            filter_stats: Dict with filter statistics from _filter_excavation_tickets
        """
        stats = {
            'Metric': [],
            'Value': [],
            'Notes': []
        }

        total_excavation_tickets = len(tickets_df)
        assigned_tickets = len(tickets_df[tickets_df['route_leg'] != 'Unassigned'])

        # Add filter statistics if available
        if filter_stats:
            stats['Metric'].append('Total All Tickets')
            stats['Value'].append(filter_stats['total_tickets'])
            stats['Notes'].append('All geocoded tickets (including non-excavation)')

            stats['Metric'].append('Excavation Tickets')
            stats['Value'].append(filter_stats['excavation_tickets'])
            stats['Notes'].append('Normal, Emergency, and DigUp tickets only')

            if filter_stats['excluded_tickets'] > 0:
                stats['Metric'].append('Excluded Tickets')
                stats['Value'].append(filter_stats['excluded_tickets'])
                excluded_types = ', '.join([f"{k}: {v}" for k, v in sorted(filter_stats['excluded_by_type'].items())])
                stats['Notes'].append(f"Non-excavation tickets ({excluded_types})")

        stats['Metric'].extend([
            'Assigned to Route Legs',
            'Unassigned (>500m from route)',
            'Assignment Rate',
        ])
        stats['Value'].extend([
            assigned_tickets,
            total_excavation_tickets - assigned_tickets,
            assigned_tickets / total_excavation_tickets if total_excavation_tickets > 0 else 0
        ])
        stats['Notes'].extend([
            'Excavation tickets within 500m of route legs',
            'Excavation tickets outside buffer zone',
            'Percentage of excavation tickets assigned to legs'
        ])

        return pd.DataFrame(stats)

    def _generate_leg_details(self, tickets_df: pd.DataFrame, years_span: float = 1.0) -> pd.DataFrame:
        """Generate detailed statistics per route leg with annualization.

        Args:
            tickets_df: DataFrame with tickets assigned to legs
            years_span: Number of years data spans for annualization
        """
        # Group by route leg
        leg_groups = tickets_df.groupby('route_leg')

        details = []
        for leg_name, group in leg_groups:
            if leg_name is None or leg_name == 'Unassigned':
                continue

            # Get leg geometry to calculate length
            leg_row = self.route_legs[self.route_legs['name'] == leg_name]
            if not leg_row.empty:
                # Calculate length in miles (convert from degrees to meters to miles)
                leg_geom = leg_row.iloc[0].geometry
                # Reproject to UTM for accurate length measurement
                leg_utm = gpd.GeoSeries([leg_geom], crs='EPSG:4326').to_crs('EPSG:32613')  # UTM Zone 13N for Texas
                leg_length_m = leg_utm.length.iloc[0]
                leg_length_mi = leg_length_m / 1609.34
            else:
                leg_length_mi = None

            # Basic counts (total in dataset)
            total_tickets_dataset = len(group)

            # Annualized count
            total_tickets = round(total_tickets_dataset / years_span)

            # By ticket type (if column exists) - also annualized
            if 'ticket_type' in group.columns:
                emergency_tickets_dataset = len(group[group['ticket_type'] == 'Emergency'])
                normal_tickets_dataset = len(group[group['ticket_type'] != 'Emergency'])
                emergency_tickets = round(emergency_tickets_dataset / years_span)
                normal_tickets = round(normal_tickets_dataset / years_span)
            else:
                emergency_tickets = 0
                normal_tickets = total_tickets

            # By duration (if column exists)
            if 'duration' in group.columns:
                duration_counts = group['duration'].value_counts().to_dict()
            else:
                duration_counts = {}

            # By work type (if column exists)
            if 'work_type' in group.columns:
                work_type_counts = group['work_type'].value_counts()
                top_work_type = work_type_counts.index[0] if len(work_type_counts) > 0 else 'N/A'
            else:
                top_work_type = 'N/A'

            # Quality metrics
            avg_confidence = group['confidence'].mean() if 'confidence' in group.columns else None

            # Calculate tickets per mile
            tickets_per_mile = total_tickets / leg_length_mi if leg_length_mi and leg_length_mi > 0 else None

            details.append({
                'Route Leg': leg_name,
                'Leg Length (mi)': round(leg_length_mi, 2) if leg_length_mi else 'N/A',
                'Total Tickets': total_tickets,
                'Tickets/Mile': round(tickets_per_mile, 2) if tickets_per_mile else 'N/A',
                'Emergency Tickets': emergency_tickets,
                'Normal Tickets': normal_tickets,
                'Emergency %': emergency_tickets/total_tickets if total_tickets > 0 else 0,
                '1-3 Day Duration': duration_counts.get('1-3 Days', 0) + duration_counts.get('1 Day', 0),
                '4-10 Day Duration': duration_counts.get('4-10 Days', 0),
                '11+ Day Duration': duration_counts.get('11+ Days', 0),
                'Top Work Type': top_work_type,
                'Avg Confidence': avg_confidence if avg_confidence is not None else 0,
            })

        return pd.DataFrame(details)

    def _generate_breakdowns(self, tickets_df: pd.DataFrame) -> pd.DataFrame:
        """Generate detailed breakdowns by various dimensions."""
        breakdowns = []

        # For each leg, generate breakdowns
        for leg_name in tickets_df['route_leg'].unique():
            if leg_name is None or leg_name == 'Unassigned':
                continue

            leg_tickets = tickets_df[tickets_df['route_leg'] == leg_name]

            # By ticket type (if column exists)
            if 'ticket_type' in leg_tickets.columns:
                for ticket_type in leg_tickets['ticket_type'].unique():
                    if pd.isna(ticket_type):
                        continue
                    count = len(leg_tickets[leg_tickets['ticket_type'] == ticket_type])
                    breakdowns.append({
                        'Route Leg': leg_name,
                        'Category': 'Ticket Type',
                        'Subcategory': ticket_type,
                        'Count': count,
                        'Percentage': count/len(leg_tickets) if len(leg_tickets) > 0 else 0
                    })

            # By duration (if column exists)
            if 'duration' in leg_tickets.columns:
                for duration in leg_tickets['duration'].unique():
                    if pd.isna(duration):
                        continue
                    count = len(leg_tickets[leg_tickets['duration'] == duration])
                    breakdowns.append({
                        'Route Leg': leg_name,
                        'Category': 'Duration',
                        'Subcategory': duration,
                        'Count': count,
                        'Percentage': count/len(leg_tickets) if len(leg_tickets) > 0 else 0
                    })

            # By work type (top 5) (if column exists)
            if 'work_type' in leg_tickets.columns:
                work_type_counts = leg_tickets['work_type'].value_counts().head(5)
                for work_type, count in work_type_counts.items():
                    if pd.isna(work_type):
                        continue
                    breakdowns.append({
                        'Route Leg': leg_name,
                        'Category': 'Work Type',
                        'Subcategory': work_type,
                        'Count': count,
                        'Percentage': count/len(leg_tickets) if len(leg_tickets) > 0 else 0
                    })

        return pd.DataFrame(breakdowns)

    def _generate_cost_projections(self, tickets_df: pd.DataFrame, leg_details_df: pd.DataFrame) -> pd.DataFrame:
        """Generate cost projection estimates per leg.

        Returns DataFrame with ticket counts - formulas will be added during Excel write.
        """
        projections = []

        for _, leg in leg_details_df.iterrows():
            leg_name = leg['Route Leg']
            total_tickets = leg['Total Tickets']

            projections.append({
                'Route Leg': leg_name,
                'Annual Tickets': total_tickets,
            })

        return pd.DataFrame(projections)

    def _write_inputs_sheet(
        self,
        writer: pd.ExcelWriter,
        leg_details_df: pd.DataFrame,
        filter_stats: dict = None,
    ) -> None:
        """Write Inputs sheet with all parameters centralized."""
        from openpyxl.styles import Font

        workbook = writer.book
        worksheet = workbook.create_sheet('Inputs', 1)  # Insert as second sheet

        # Row 1: Column headers
        worksheet['A1'] = "Input Parameter"
        worksheet['A1'].font = Font(name="Arial", bold=True)
        worksheet['B1'] = "Value"
        worksheet['B1'].font = Font(name="Arial", bold=True)
        worksheet['C1'] = "Source/Notes"
        worksheet['C1'].font = Font(name="Arial", bold=True)
        worksheet['D1'] = "Referenced From"
        worksheet['D1'].font = Font(name="Arial", bold=True)
        worksheet['E1'] = "Source"
        worksheet['E1'].font = Font(name="Arial", bold=True)

        # Row 2: Locate Fee Per Ticket
        worksheet['A2'] = "Locate Fee Per Ticket"
        worksheet['B2'] = 25.00
        worksheet['B2'].number_format = '"$"#,##0.00'
        worksheet['C2'] = "How much does a locate cost?"
        worksheet['D2'] = "Cost Projections"

        # Row 3: Average Cost Per Strike Repair
        worksheet['A3'] = "Average Cost Per Strike Repair"
        worksheet['B3'] = 11728.00
        worksheet['B3'].number_format = '"$"#,##0.00'
        worksheet['C3'] = "What does it cost, on average to fix a fiber cut?"
        worksheet['D3'] = "Cost Projections"

        # Row 4: Probability of Damage
        worksheet['A4'] = "Probability of Damage"
        worksheet['B4'] = 0.0056
        worksheet['B4'].number_format = '0.00%'
        worksheet['C4'] = "What is the average number of tickets that result in damage?"
        worksheet['D4'] = "Cost Projections"

        # Row 5: Probability of Damage Telcom
        worksheet['A5'] = "Probability of Damage Telcom"
        worksheet['B5'] = 0.47
        worksheet['B5'].number_format = '0%'
        worksheet['C5'] = "What portion of tickets that result in damage are damager to"
        worksheet['D5'] = "Cost Projections"

        # Row 6: Expected Probability of Damage (formula)
        worksheet['A6'] = "Expected Probability of Damage"
        worksheet['B6'] = '=B4*B5'
        worksheet['B6'].number_format = '0.00%'
        worksheet['C6'] = "Probability given the two statistics above"
        worksheet['D6'] = "Cost Projections"

        # Row 7: Margin
        worksheet['A7'] = "Margin"
        worksheet['B7'] = 0.01  # 1%
        worksheet['B7'].number_format = '0.00%'
        worksheet['C7'] = "What margin do I want to build into estimages?"
        worksheet['D7'] = "Maintenance Estimate"

        # Row 8: Buffer Distance
        worksheet['A8'] = "Buffer Distance (Meters)"
        worksheet['B8'] = self.buffer_distance_m
        worksheet['C8'] = "Within what distance should we include tickets?"
        worksheet['E8'] = "Geocoding pipeline configuration"

        # Row 9: Insurance Rate
        worksheet['A9'] = "Insurance Rate"
        worksheet['B9'] = 0.0035
        worksheet['C9'] = "What does the Property and Business Insurance cost per insta"
        worksheet['D9'] = "Maintenance Estimate"

        # Row 10: Build Cost
        worksheet['A10'] = "Build Cost"
        worksheet['B10'] = 3500000
        worksheet['B10'].number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
        worksheet['C10'] = "How much did we charge the client to build the plant?"
        worksheet['E10'] = "Build cost estimate"

        # Row 11: Per Mile Up-front Payment (Percentage)
        worksheet['A11'] = "Per Mile Up-front Payment (Percentage)"
        worksheet['B11'] = 0.05  # 5%
        worksheet['B11'].number_format = '0%'
        worksheet['C11'] = "What percentage of the build cost should we ask for up front"
        worksheet['D11'] = "Up-Front Payment (B12)"
        worksheet['E11'] = "Industry is 5%-15%"

        # Row 12: Up-Front Payment
        worksheet['A12'] = "Up-Front Payment"
        worksheet['B12'] = '=B10*B11'
        worksheet['B12'].number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
        worksheet['C12'] = "What is the non-recurring cost of the IRU?"

        # Set column widths to match reference
        worksheet.column_dimensions['A'].width = 39.57
        worksheet.column_dimensions['B'].width = 16.43
        worksheet.column_dimensions['C'].width = 43.0
        worksheet.column_dimensions['D'].width = 24.86
        worksheet.column_dimensions['E'].width = 33.14
        worksheet.column_dimensions['F'].width = 8.86

        # Apply Arial font to all cells
        self._apply_arial_font(worksheet, max_row=12, max_col=6)

    def _write_summary_sheet(
        self,
        writer: pd.ExcelWriter,
        summary_df: pd.DataFrame,
        leg_details_df: pd.DataFrame,
        project_name: str
    ) -> dict:
        """Write formatted summary sheet with pricing table.

        Returns:
            Dictionary with leg_row_mapping, total_cost_row, and total_cost_lease_row for Quote sheet references
        """
        from openpyxl.styles import Font

        workbook = writer.book
        worksheet = workbook.create_sheet('Maintenance Estimate', 0)

        # Row 1: Title
        worksheet['A1'] = "Project - O&M Maintenance Estimate"
        worksheet['A1'].font = Font(name="Arial", bold=True)

        # Row 2: Generated timestamp
        worksheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Row 3: Blank
        # Row 4: Statistics header
        worksheet['A4'] = "Statistics"
        worksheet['A4'].font = Font(name="Arial", bold=True)
        worksheet['B4'] = "Counts"
        worksheet['C4'] = "Description"

        # Rows 5+: Summary statistics (dynamic based on filter stats)
        row = 5
        for _, stat_row in summary_df.iterrows():
            worksheet[f'A{row}'] = stat_row['Metric']
            value = stat_row['Value']

            # Handle different value types
            if 'Rate' in stat_row['Metric'] and isinstance(value, (int, float)):
                # Skip row 10, add blank row
                if row == 10:
                    row += 1
                # Assignment Rate - numeric value with percentage format
                worksheet[f'B{row}'] = value
                worksheet[f'B{row}'].number_format = '0.0%'
            else:
                # All other values - store as-is
                worksheet[f'B{row}'] = value

            worksheet[f'C{row}'] = stat_row['Notes']
            row += 1

        # Dynamic positioning based on number of summary stats
        current_row = row + 2  # Blank row after summary stats

        # TICKETS PER ROUTE LEG section
        current_row += 1
        worksheet[f'A{current_row}'] = "Route Leg"
        worksheet[f'A{current_row}'].font = Font(name="Arial", bold=True)
        worksheet[f'B{current_row}'] = "Total Tickets"
        worksheet[f'C{current_row}'] = "Leg Length (Miles)"
        worksheet[f'D{current_row}'] = "Tickets/Mile"

        # Route leg data rows
        leg_start_row = current_row + 1
        leg_row_mapping = {}  # Map leg names to row numbers for Leg Details sheet
        for _, leg in leg_details_df.iterrows():
            current_row += 1
            leg_name = leg['Route Leg']
            worksheet[f'A{current_row}'] = leg_name
            worksheet[f'B{current_row}'] = leg['Total Tickets']
            worksheet[f'C{current_row}'] = leg['Leg Length (mi)'] if isinstance(leg['Leg Length (mi)'], (int, float)) else 0
            worksheet[f'D{current_row}'] = leg['Tickets/Mile'] if isinstance(leg['Tickets/Mile'], (int, float)) else 0
            leg_row_mapping[leg_name] = current_row  # Store for Quote sheet
        leg_end_row = current_row

        # Blank rows + NOC section
        current_row += 3
        worksheet[f'A{current_row}'] = "Item"
        worksheet[f'A{current_row}'].font = Font(name="Arial", bold=True)
        worksheet[f'B{current_row}'] = "Cost"
        worksheet[f'B{current_row}'].font = Font(name="Arial", bold=True)
        worksheet[f'C{current_row}'] = "Type"
        worksheet[f'C{current_row}'].font = Font(name="Arial", bold=True)

        current_row += 1
        worksheet[f'A{current_row}'] = "NOC Up-Front"
        worksheet[f'B{current_row}'] = '=120*150'
        worksheet[f'C{current_row}'] = "Non-recurring"

        current_row += 1
        noc_monthly_row = current_row
        worksheet[f'A{current_row}'] = "Monthly NOC Cost"
        worksheet[f'B{current_row}'] = 3000.00
        worksheet[f'C{current_row}'] = "Monthly recurring"

        # Blank rows + Pricing table
        current_row += 3
        pricing_header_row = current_row
        worksheet[f'A{current_row}'] = "Line Item"
        worksheet[f'A{current_row}'].font = Font(name="Arial", bold=True)
        worksheet[f'B{current_row}'] = "Estimate"
        worksheet[f'B{current_row}'].font = Font(name="Arial", bold=True)
        worksheet[f'C{current_row}'] = "Margin"
        worksheet[f'C{current_row}'].font = Font(name="Arial", bold=True)
        worksheet[f'D{current_row}'] = "Price"
        worksheet[f'D{current_row}'].font = Font(name="Arial", bold=True)

        # NOC Monitoring Cost
        current_row += 1
        noc_row = current_row
        worksheet[f'A{current_row}'] = "NOC Monitoring Cost"
        worksheet[f'B{current_row}'] = f'=B{noc_monthly_row}*12'
        worksheet[f'B{current_row}'].number_format = '"$"#,##0.00'
        worksheet[f'C{current_row}'] = 0.15  # 15% margin
        worksheet[f'C{current_row}'].number_format = '0%'
        worksheet[f'D{current_row}'] = f'=B{current_row}*(1+C{current_row})'
        worksheet[f'D{current_row}'].number_format = '"$"#,##0.00'

        # Insurance Cost
        current_row += 1
        ins_row = current_row
        worksheet[f'A{current_row}'] = "Insurance Cost"
        worksheet[f'B{current_row}'] = '=Inputs!B10*Inputs!B9'
        worksheet[f'B{current_row}'].number_format = '"$"#,##0.00'
        worksheet[f'C{current_row}'] = 0.50  # 50% margin
        worksheet[f'C{current_row}'].number_format = '0%'
        worksheet[f'D{current_row}'] = f'=B{current_row}*(1+C{current_row})'
        worksheet[f'D{current_row}'].number_format = '"$"#,##0.00'

        # Maintenance Cost
        current_row += 1
        maint_row = current_row
        worksheet[f'A{current_row}'] = "Maintenance Cost"
        worksheet[f'B{current_row}'] = "='Cost Projections'!I6"
        worksheet[f'B{current_row}'].number_format = '"$"#,##0.00'
        worksheet[f'C{current_row}'] = 0.40  # 40% margin
        worksheet[f'C{current_row}'].number_format = '0%'
        worksheet[f'D{current_row}'] = f'=B{current_row}*(1+C{current_row})'
        worksheet[f'D{current_row}'].number_format = '"$"#,##0.00'

        # Total Annual O&M Cost
        current_row += 1
        total_row = current_row
        worksheet[f'A{current_row}'] = "Total Annual O&M Cost"
        worksheet[f'B{current_row}'] = f'=SUM(B{noc_row}:B{maint_row})'
        worksheet[f'B{current_row}'].number_format = '"$"#,##0.00'
        worksheet[f'B{current_row}'].font = Font(name="Arial", bold=True)
        worksheet[f'C{current_row}'] = "Total Price (Annual)"
        worksheet[f'C{current_row}'].font = Font(name="Arial", bold=True)
        worksheet[f'D{current_row}'] = f'=SUM(D{noc_row}:D{maint_row})'
        worksheet[f'D{current_row}'].number_format = '"$"#,##0.00'
        worksheet[f'D{current_row}'].font = Font(name="Arial", bold=True)

        # Total Cost of Lease (TCL)
        current_row += 1
        total_cost_lease_row = current_row
        worksheet[f'C{current_row}'] = "Total Cost of Lease (TCL)"
        worksheet[f'C{current_row}'].font = Font(name="Arial", bold=True)
        worksheet[f'D{current_row}'] = f'=20*D{total_row}'
        worksheet[f'D{current_row}'].number_format = '"$"#,##0.00'
        worksheet[f'D{current_row}'].font = Font(name="Arial", bold=True)

        # TCL as a Percentage of Build Cost
        current_row += 1
        worksheet[f'C{current_row}'] = "TCL as a Percentage of Build Cost"
        worksheet[f'C{current_row}'].font = Font(name="Arial", bold=True)
        worksheet[f'D{current_row}'] = f'=D{total_cost_lease_row}/Inputs!B10'
        worksheet[f'D{current_row}'].number_format = '0.0%'
        worksheet[f'D{current_row}'].font = Font(name="Arial", bold=True)

        # Store leg_start_row for Inputs sheet formula
        self._leg_start_row = leg_start_row
        self._leg_end_row = leg_end_row

        # Set column widths to match reference
        worksheet.column_dimensions['A'].width = 33.57
        worksheet.column_dimensions['B'].width = 20.14
        worksheet.column_dimensions['C'].width = 38.14
        worksheet.column_dimensions['D'].width = 19.43
        worksheet.column_dimensions['E'].width = 8.86

        # Apply Arial font to all cells
        self._apply_arial_font(worksheet, max_row=current_row, max_col=5)

        # Return mapping for Quote sheet
        return {
            'leg_row_mapping': leg_row_mapping,
            'total_cost_row': total_row,
            'total_cost_lease_row': total_cost_lease_row,
        }

    def _write_cost_projections_sheet(
        self,
        writer: pd.ExcelWriter,
        cost_projections_df: pd.DataFrame,
    ) -> None:
        """Write Cost Projections sheet with formulas referencing Inputs sheet."""
        from openpyxl.styles import Font

        workbook = writer.book
        worksheet = workbook.create_sheet('Cost Projections')

        # Row 1: Column headers
        headers = [
            "Route Leg",
            "Annual Tickets",
            "Monthly Avg Tickets",
            "Locate Fee",  # Changed from "Locate Fee ($/ticket)"
            "Annual Locate Cost",
            "Annual Adjusted Ticket Count: Probability of Damage",
            "Annual Estimated Maintenance Cost",
            "Annual Estimated Locate Cost",
            "Annual Total Maintenance Cost"
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(name="Arial", bold=True)

        # Rows 2-5: Data rows with formulas
        row = 2
        for _, leg_data in cost_projections_df.iterrows():
            # Column A: Route Leg name
            worksheet[f'A{row}'] = leg_data['Route Leg']

            # Column B: Annual Tickets (static value)
            worksheet[f'B{row}'] = leg_data['Annual Tickets']
            worksheet[f'B{row}'].number_format = '0.00'

            # Column C: Monthly Avg Tickets (formula: =B2/12)
            worksheet[f'C{row}'] = f'=B{row}/12'
            worksheet[f'C{row}'].number_format = '0.00'

            # Column D: Locate Fee (reference to Inputs!B8)
            worksheet[f'D{row}'] = "=Inputs!$B$8"
            worksheet[f'D{row}'].number_format = '"$"#,##0.00'

            # Column E: Annual Locate Cost (formula: =B2*D2)
            worksheet[f'E{row}'] = f'=B{row}*D{row}'
            worksheet[f'E{row}'].number_format = '"$"#,##0.00'

            # Column F: Annual Adjusted Ticket Count (formula: =B2*Inputs!$B$12)
            worksheet[f'F{row}'] = f"=B{row}*Inputs!$B$12"
            worksheet[f'F{row}'].number_format = '0.00'

            # Column G: Annual Estimated Maintenance Cost (formula: =F2*Inputs!$B$9)
            worksheet[f'G{row}'] = f"=F{row}*Inputs!$B$9"
            worksheet[f'G{row}'].number_format = '"$"#,##0.00'

            # Column H: Annual Estimated Locate Cost (formula: =E2)
            worksheet[f'H{row}'] = f'=E{row}'
            worksheet[f'H{row}'].number_format = '"$"#,##0.00'

            # Column I: Annual Total Maintenance Cost (formula: =G2+H2)
            worksheet[f'I{row}'] = f'=G{row}+H{row}'
            worksheet[f'I{row}'].number_format = '"$"#,##0.00'

            row += 1

        # Row 6: Totals row
        last_data_row = row - 1
        worksheet[f'A{row}'] = "TOTAL"
        worksheet[f'A{row}'].font = Font(name="Arial", bold=True)

        # Column B: Total annual tickets
        worksheet[f'B{row}'] = f'=SUM(B2:B{last_data_row})'
        worksheet[f'B{row}'].number_format = '0.00'
        worksheet[f'B{row}'].font = Font(name="Arial", bold=True)

        # Column C: Total monthly avg tickets
        worksheet[f'C{row}'] = f'=SUM(C2:C{last_data_row})'
        worksheet[f'C{row}'].number_format = '0.00'
        worksheet[f'C{row}'].font = Font(name="Arial", bold=True)

        # Column D: Blank (no total for unit price)
        # Column E: Total annual locate cost
        worksheet[f'E{row}'] = f'=SUM(E2:E{last_data_row})'
        worksheet[f'E{row}'].number_format = '"$"#,##0.00'
        worksheet[f'E{row}'].font = Font(name="Arial", bold=True)

        # Column F: Total adjusted ticket count
        worksheet[f'F{row}'] = f'=SUM(F2:F{last_data_row})'
        worksheet[f'F{row}'].number_format = '0.00'
        worksheet[f'F{row}'].font = Font(name="Arial", bold=True)

        # Column G: Total estimated maintenance cost
        worksheet[f'G{row}'] = f'=SUM(G2:G{last_data_row})'
        worksheet[f'G{row}'].number_format = '"$"#,##0.00'
        worksheet[f'G{row}'].font = Font(name="Arial", bold=True)

        # Column H: Total estimated locate cost
        worksheet[f'H{row}'] = f'=SUM(H2:H{last_data_row})'
        worksheet[f'H{row}'].number_format = '"$"#,##0.00'
        worksheet[f'H{row}'].font = Font(name="Arial", bold=True)

        # Column I: Total maintenance cost (THIS IS REFERENCED IN Maintenance Estimate B21)
        worksheet[f'I{row}'] = f'=SUM(I2:I{last_data_row})'
        worksheet[f'I{row}'].number_format = '"$"#,##0.00'
        worksheet[f'I{row}'].font = Font(name='Arial', bold=True)

        # Apply Arial font to all cells
        self._apply_arial_font(worksheet, max_row=row, max_col=9)


def generate_maintenance_estimate(
    tickets_df: pd.DataFrame,
    kmz_path: Path,
    output_path: Path,
    project_name: str = "Project",
    buffer_distance_m: float = 500.0,
) -> None:
    """Convenience function to generate maintenance estimate.

    Args:
        tickets_df: DataFrame with geocoded tickets
        kmz_path: Path to KMZ file with route legs
        output_path: Path for output Excel file
        project_name: Name of project
        buffer_distance_m: Buffer distance for ticket assignment
    """
    generator = MaintenanceEstimateGenerator(
        kmz_path=kmz_path,
        buffer_distance_m=buffer_distance_m
    )

    generator.generate_estimate(
        tickets_df=tickets_df,
        output_path=output_path,
        project_name=project_name
    )


if __name__ == "__main__":
    # Test the maintenance estimate generator
    print("MaintenanceEstimateGenerator - Test Mode\n")

    kmz_path = Path("../projects/wink/route/wink.kmz")

    if kmz_path.exists():
        print(f"Loading route from {kmz_path}")
        generator = MaintenanceEstimateGenerator(kmz_path=kmz_path)

        print(f"\n✓ Loaded {len(generator.route_legs)} route legs:")
        for idx, leg in generator.route_legs.iterrows():
            print(f"  - {leg['name']}")
    else:
        print(f"Test file not found: {kmz_path}")
