"""
Maintenance Estimate Generator

Generates O&M maintenance cost estimates by analyzing ticket distribution
across route legs/segments. Outputs an Excel workbook with:
- Tickets per leg/segment
- Ticket breakdowns by type, duration, work type
- Cost calculations (locate, maintenance, etc.)
- Summary statistics
"""

import logging
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional
from datetime import datetime
from zipfile import ZipFile
import xml.etree.ElementTree as ET

import geopandas as gpd
import pandas as pd
from shapely.geometry import Point, LineString
from shapely.ops import nearest_points

logger = logging.getLogger(__name__)


class MaintenanceEstimateGenerator:
    """Generates maintenance cost estimates by route leg."""

    def __init__(
        self,
        kmz_path: Path,
        buffer_distance_m: float = 500.0,
    ):
        """Initialize maintenance estimate generator.

        Args:
            kmz_path: Path to KMZ file with route legs
            buffer_distance_m: Buffer distance for assigning tickets to legs (default 500m)
        """
        self.kmz_path = kmz_path
        self.buffer_distance_m = buffer_distance_m

        # Load route legs
        self.route_legs = self._load_route_legs()

        logger.info(f"Loaded {len(self.route_legs)} route legs from {kmz_path}")

    def _load_route_legs(self) -> gpd.GeoDataFrame:
        """Load route legs from KMZ file.

        Returns:
            GeoDataFrame with route leg geometries and names
        """
        try:
            # Extract KML from KMZ
            with ZipFile(self.kmz_path, 'r') as zf:
                kml_files = [name for name in zf.namelist() if name.endswith('.kml')]
                if not kml_files:
                    raise ValueError(f"No KML file found in {self.kmz_path}")

                kml_content = zf.read(kml_files[0])

            # Parse KML to extract route legs
            root = ET.fromstring(kml_content)
            ns = {'kml': 'http://www.opengis.net/kml/2.2'}

            placemarks = root.findall('.//kml:Placemark', ns)

            legs = []
            for placemark in placemarks:
                name_elem = placemark.find('kml:name', ns)
                linestring_elem = placemark.find('.//kml:LineString', ns)

                if linestring_elem is not None:
                    coords_elem = linestring_elem.find('kml:coordinates', ns)
                    if coords_elem is not None:
                        # Parse coordinates (format: "lon,lat,alt lon,lat,alt ...")
                        coords_text = coords_elem.text.strip()
                        coords = []
                        for coord_str in coords_text.split():
                            parts = coord_str.split(',')
                            if len(parts) >= 2:
                                lon, lat = float(parts[0]), float(parts[1])
                                coords.append((lon, lat))

                        if coords:
                            name = name_elem.text if name_elem is not None else "Unnamed"
                            legs.append({
                                'name': name,
                                'geometry': LineString(coords)
                            })

            # Create GeoDataFrame
            gdf = gpd.GeoDataFrame(legs, crs='EPSG:4326')
            return gdf

        except Exception as e:
            logger.error(f"Failed to load route legs: {e}")
            raise

    def assign_tickets_to_legs(
        self,
        tickets_df: pd.DataFrame,
    ) -> pd.DataFrame:
        """Assign tickets to route legs based on proximity.

        Args:
            tickets_df: DataFrame with geocoded tickets
                Required columns: ticket_number, latitude, longitude

        Returns:
            DataFrame with added columns:
                - route_leg: Name of assigned route leg
                - distance_to_leg_m: Distance to nearest point on leg
        """
        # Convert tickets to GeoDataFrame
        tickets_gdf = gpd.GeoDataFrame(
            tickets_df,
            geometry=gpd.points_from_xy(tickets_df.longitude, tickets_df.latitude),
            crs='EPSG:4326'
        )

        # For each ticket, find nearest leg
        assignments = []

        for idx, ticket in tickets_gdf.iterrows():
            if ticket.geometry is None or pd.isna(ticket.latitude):
                assignments.append({
                    'route_leg': None,
                    'distance_to_leg_m': None
                })
                continue

            # Find nearest leg
            min_distance = float('inf')
            nearest_leg = None

            for leg_idx, leg in self.route_legs.iterrows():
                # Calculate distance (in degrees, then convert to meters)
                nearest_pt_on_leg, nearest_pt_from_ticket = nearest_points(
                    leg.geometry, ticket.geometry
                )

                distance_m = self._calculate_distance_meters(
                    ticket.latitude, ticket.longitude,
                    nearest_pt_on_leg.y, nearest_pt_on_leg.x
                )

                if distance_m < min_distance:
                    min_distance = distance_m
                    nearest_leg = leg['name']

            # Only assign if within buffer distance
            if min_distance <= self.buffer_distance_m:
                assignments.append({
                    'route_leg': nearest_leg,
                    'distance_to_leg_m': round(min_distance, 2)
                })
            else:
                assignments.append({
                    'route_leg': 'Unassigned',
                    'distance_to_leg_m': round(min_distance, 2)
                })

        # Add assignments to dataframe
        result_df = tickets_df.copy()
        result_df['route_leg'] = [a['route_leg'] for a in assignments]
        result_df['distance_to_leg_m'] = [a['distance_to_leg_m'] for a in assignments]

        return result_df

    def _calculate_distance_meters(
        self,
        lat1: float,
        lng1: float,
        lat2: float,
        lng2: float,
    ) -> float:
        """Calculate distance in meters between two lat/lng points using Haversine."""
        from math import radians, sin, cos, sqrt, atan2

        R = 6371000  # Earth radius in meters

        lat1_rad = radians(lat1)
        lat2_rad = radians(lat2)
        dlat = radians(lat2 - lat1)
        dlng = radians(lng2 - lng1)

        a = sin(dlat / 2) ** 2 + cos(lat1_rad) * cos(lat2_rad) * sin(dlng / 2) ** 2
        c = 2 * atan2(sqrt(a), sqrt(1 - a))

        return R * c

    def generate_estimate(
        self,
        tickets_df: pd.DataFrame,
        output_path: Path,
        project_name: str = "Project",
    ) -> None:
        """Generate maintenance estimate Excel workbook.

        Args:
            tickets_df: DataFrame with geocoded tickets
            output_path: Path for output Excel file
            project_name: Name of project for report title
        """
        logger.info(f"Generating maintenance estimate for {len(tickets_df)} tickets")

        # Assign tickets to legs
        tickets_with_legs = self.assign_tickets_to_legs(tickets_df)

        # Generate statistics
        summary_df = self._generate_summary_stats(tickets_with_legs)
        leg_details_df = self._generate_leg_details(tickets_with_legs)
        breakdown_df = self._generate_breakdowns(tickets_with_legs)
        cost_projections_df = self._generate_cost_projections(tickets_with_legs, leg_details_df)

        # Create Excel workbook
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Summary with estimate inputs
            self._write_summary_sheet(writer, summary_df, leg_details_df, project_name)

            # Sheet 2: Leg Details
            leg_details_df.to_excel(writer, sheet_name='Leg Details', index=False)

            # Sheet 3: Cost Projections
            cost_projections_df.to_excel(writer, sheet_name='Cost Projections', index=False)

            # Sheet 4: Breakdowns
            breakdown_df.to_excel(writer, sheet_name='Ticket Breakdowns', index=False)

            # Sheet 5: Raw Data (tickets with leg assignments)
            tickets_with_legs.to_excel(writer, sheet_name='Ticket Assignments', index=False)

        logger.info(f"Maintenance estimate saved to {output_path}")

    def _generate_summary_stats(self, tickets_df: pd.DataFrame) -> pd.DataFrame:
        """Generate overall summary statistics."""
        stats = {
            'Metric': [],
            'Value': [],
            'Notes': []
        }

        total_tickets = len(tickets_df)
        assigned_tickets = len(tickets_df[tickets_df['route_leg'] != 'Unassigned'])

        stats['Metric'].extend([
            'Total Tickets',
            'Assigned to Route Legs',
            'Unassigned (>500m from route)',
            'Assignment Rate',
        ])
        stats['Value'].extend([
            total_tickets,
            assigned_tickets,
            total_tickets - assigned_tickets,
            f"{(assigned_tickets/total_tickets*100):.1f}%" if total_tickets > 0 else "N/A"
        ])
        stats['Notes'].extend([
            'All geocoded tickets',
            'Within 500m of route legs',
            'Outside buffer zone',
            'Percentage of tickets assigned to legs'
        ])

        return pd.DataFrame(stats)

    def _generate_leg_details(self, tickets_df: pd.DataFrame) -> pd.DataFrame:
        """Generate detailed statistics per route leg."""
        # Group by route leg
        leg_groups = tickets_df.groupby('route_leg')

        details = []
        for leg_name, group in leg_groups:
            if leg_name is None or leg_name == 'Unassigned':
                continue

            # Get leg geometry to calculate length
            leg_row = self.route_legs[self.route_legs['name'] == leg_name]
            if not leg_row.empty:
                # Calculate length in miles (convert from degrees to meters to miles)
                leg_geom = leg_row.iloc[0].geometry
                # Reproject to UTM for accurate length measurement
                leg_utm = gpd.GeoSeries([leg_geom], crs='EPSG:4326').to_crs('EPSG:32613')  # UTM Zone 13N for Texas
                leg_length_m = leg_utm.length.iloc[0]
                leg_length_mi = leg_length_m / 1609.34
            else:
                leg_length_mi = None

            # Basic counts
            total_tickets = len(group)

            # By ticket type
            emergency_tickets = len(group[group['ticket_type'] == 'Emergency'])
            normal_tickets = len(group[group['ticket_type'] != 'Emergency'])

            # By duration
            duration_counts = group['duration'].value_counts().to_dict()

            # By work type
            work_type_counts = group['work_type'].value_counts()
            top_work_type = work_type_counts.index[0] if len(work_type_counts) > 0 else 'N/A'

            # Quality metrics
            avg_confidence = group['confidence'].mean() if 'confidence' in group.columns else None

            # Calculate tickets per mile
            tickets_per_mile = total_tickets / leg_length_mi if leg_length_mi and leg_length_mi > 0 else None

            details.append({
                'Route Leg': leg_name,
                'Leg Length (mi)': round(leg_length_mi, 2) if leg_length_mi else 'N/A',
                'Total Tickets': total_tickets,
                'Tickets/Mile': round(tickets_per_mile, 2) if tickets_per_mile else 'N/A',
                'Emergency Tickets': emergency_tickets,
                'Normal Tickets': normal_tickets,
                'Emergency %': f"{(emergency_tickets/total_tickets*100):.1f}%" if total_tickets > 0 else "0%",
                '1-3 Day Duration': duration_counts.get('1-3 Days', 0) + duration_counts.get('1 Day', 0),
                '4-10 Day Duration': duration_counts.get('4-10 Days', 0),
                '11+ Day Duration': duration_counts.get('11+ Days', 0),
                'Top Work Type': top_work_type,
                'Avg Confidence': f"{avg_confidence:.2%}" if avg_confidence is not None else 'N/A',
            })

        return pd.DataFrame(details)

    def _generate_breakdowns(self, tickets_df: pd.DataFrame) -> pd.DataFrame:
        """Generate detailed breakdowns by various dimensions."""
        breakdowns = []

        # For each leg, generate breakdowns
        for leg_name in tickets_df['route_leg'].unique():
            if leg_name is None or leg_name == 'Unassigned':
                continue

            leg_tickets = tickets_df[tickets_df['route_leg'] == leg_name]

            # By ticket type
            for ticket_type in leg_tickets['ticket_type'].unique():
                if pd.isna(ticket_type):
                    continue
                count = len(leg_tickets[leg_tickets['ticket_type'] == ticket_type])
                breakdowns.append({
                    'Route Leg': leg_name,
                    'Category': 'Ticket Type',
                    'Subcategory': ticket_type,
                    'Count': count,
                    'Percentage': f"{(count/len(leg_tickets)*100):.1f}%"
                })

            # By duration
            for duration in leg_tickets['duration'].unique():
                if pd.isna(duration):
                    continue
                count = len(leg_tickets[leg_tickets['duration'] == duration])
                breakdowns.append({
                    'Route Leg': leg_name,
                    'Category': 'Duration',
                    'Subcategory': duration,
                    'Count': count,
                    'Percentage': f"{(count/len(leg_tickets)*100):.1f}%"
                })

            # By work type (top 5)
            work_type_counts = leg_tickets['work_type'].value_counts().head(5)
            for work_type, count in work_type_counts.items():
                if pd.isna(work_type):
                    continue
                breakdowns.append({
                    'Route Leg': leg_name,
                    'Category': 'Work Type',
                    'Subcategory': work_type,
                    'Count': count,
                    'Percentage': f"{(count/len(leg_tickets)*100):.1f}%"
                })

        return pd.DataFrame(breakdowns)

    def _generate_cost_projections(self, tickets_df: pd.DataFrame, leg_details_df: pd.DataFrame) -> pd.DataFrame:
        """Generate cost projection estimates per leg."""
        projections = []

        # Default cost assumptions (can be overridden)
        DEFAULT_LOCATE_FEE = 3.50  # $ per ticket
        EMERGENCY_MULTIPLIER = 2.0  # Emergency tickets cost more

        for _, leg in leg_details_df.iterrows():
            leg_name = leg['Route Leg']
            total_tickets = leg['Total Tickets']
            emergency_tickets = leg['Emergency Tickets']
            normal_tickets = leg['Normal Tickets']

            # Calculate locate costs
            normal_locate_cost = normal_tickets * DEFAULT_LOCATE_FEE
            emergency_locate_cost = emergency_tickets * DEFAULT_LOCATE_FEE * EMERGENCY_MULTIPLIER
            total_locate_cost = normal_locate_cost + emergency_locate_cost

            # Annual projection (assume data represents one year)
            annual_tickets = total_tickets
            annual_locate_cost = total_locate_cost

            # Monthly average
            monthly_avg_tickets = annual_tickets / 12
            monthly_avg_locate_cost = annual_locate_cost / 12

            projections.append({
                'Route Leg': leg_name,
                'Annual Tickets': annual_tickets,
                'Monthly Avg Tickets': round(monthly_avg_tickets, 1),
                'Locate Fee ($/ticket)': f"${DEFAULT_LOCATE_FEE:.2f}",
                'Annual Locate Cost': f"${annual_locate_cost:,.2f}",
                'Monthly Avg Locate Cost': f"${monthly_avg_locate_cost:,.2f}",
                'Strike Repair Cost': 'USER INPUT REQUIRED',
                'Notes': 'Default locate fee applied. Update with actual rates.'
            })

        return pd.DataFrame(projections)

    def _write_summary_sheet(
        self,
        writer: pd.ExcelWriter,
        summary_df: pd.DataFrame,
        leg_details_df: pd.DataFrame,
        project_name: str
    ) -> None:
        """Write formatted summary sheet with estimate template."""
        from openpyxl.styles import Font, PatternFill, Alignment

        workbook = writer.book
        worksheet = workbook.create_sheet('Maintenance Estimate', 0)  # Insert as first sheet

        # Title section
        worksheet['A1'] = f"{project_name} - O&M Maintenance Estimate"
        worksheet['A1'].font = Font(size=14, bold=True)
        worksheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

        # Summary statistics
        row = 4
        worksheet[f'A{row}'] = "SUMMARY STATISTICS"
        worksheet[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        for _, stat in summary_df.iterrows():
            worksheet[f'A{row}'] = stat['Metric']
            worksheet[f'B{row}'] = stat['Value']
            worksheet[f'C{row}'] = stat['Notes']
            row += 1

        # Route leg summary
        row += 2
        worksheet[f'A{row}'] = "TICKETS PER ROUTE LEG"
        worksheet[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        worksheet[f'A{row}'] = "Route Leg"
        worksheet[f'B{row}'] = "Total Tickets"
        worksheet[f'C{row}'] = "Leg Length (mi)"
        worksheet[f'D{row}'] = "Tickets/Mile"
        for col in ['A', 'B', 'C', 'D']:
            worksheet[f'{col}{row}'].font = Font(bold=True)
        row += 1

        for _, leg in leg_details_df.iterrows():
            worksheet[f'A{row}'] = leg['Route Leg']
            worksheet[f'B{row}'] = leg['Total Tickets']
            worksheet[f'C{row}'] = leg['Leg Length (mi)']
            worksheet[f'D{row}'] = leg['Tickets/Mile']
            row += 1

        # Estimate inputs section
        row += 2
        worksheet[f'A{row}'] = "MAINTENANCE ESTIMATE INPUTS"
        worksheet[f'A{row}'].font = Font(bold=True, size=12)
        row += 1

        # Column headers
        worksheet[f'A{row}'] = "Input Parameter"
        worksheet[f'B{row}'] = "Value"
        worksheet[f'C{row}'] = "Unit"
        worksheet[f'D{row}'] = "Source/Notes"
        for col in ['A', 'B', 'C', 'D']:
            worksheet[f'{col}{row}'].font = Font(bold=True)
        row += 1

        # Calculate total route length
        total_route_length = leg_details_df['Leg Length (mi)'].apply(
            lambda x: float(x) if isinstance(x, (int, float)) else 0
        ).sum()

        # Calculate total annual tickets
        total_annual_tickets = leg_details_df['Total Tickets'].sum()

        # Add estimate input rows
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        inputs = [
            # Calculated inputs
            ("Total Route Length", round(total_route_length, 2), "miles", "Calculated from KMZ route legs"),
            ("Total Annual Tickets", total_annual_tickets, "tickets/year", "Based on geocoded ticket data"),
            ("Average Tickets/Mile", round(total_annual_tickets/total_route_length, 2) if total_route_length > 0 else "N/A", "tickets/mi/yr", "Calculated average across all legs"),
            ("Buffer Distance", self.buffer_distance_m, "meters", "Ticket assignment threshold"),
            ("", "", "", ""),  # Blank row

            # User input required
            ("Locate Fee Per Ticket", "", "$/ticket", "USER INPUT REQUIRED"),
            ("Average Cost Per Strike Repair", "", "$", "USER INPUT REQUIRED"),
            ("Expected Strike Rate", "", "strikes/year", "USER INPUT REQUIRED"),
            ("", "", "", ""),  # Blank row

            # NOC/Operations
            ("NOC Monitoring Cost", "", "$/month", "USER INPUT REQUIRED - Network Operations Center"),
            ("Insurance Cost", "", "$/year", "USER INPUT REQUIRED"),
            ("", "", "", ""),  # Blank row

            # Initial costs
            ("Initial Setup/Activation", "", "$", "USER INPUT REQUIRED - One-time costs"),
        ]

        for input_name, value, unit, notes in inputs:
            worksheet[f'A{row}'] = input_name
            worksheet[f'B{row}'] = value
            worksheet[f'C{row}'] = unit
            worksheet[f'D{row}'] = notes

            if "USER INPUT REQUIRED" in notes:
                worksheet[f'B{row}'].fill = yellow_fill

            row += 1

        # Set column widths
        worksheet.column_dimensions['A'].width = 35
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 50


def generate_maintenance_estimate(
    tickets_df: pd.DataFrame,
    kmz_path: Path,
    output_path: Path,
    project_name: str = "Project",
    buffer_distance_m: float = 500.0,
) -> None:
    """Convenience function to generate maintenance estimate.

    Args:
        tickets_df: DataFrame with geocoded tickets
        kmz_path: Path to KMZ file with route legs
        output_path: Path for output Excel file
        project_name: Name of project
        buffer_distance_m: Buffer distance for ticket assignment
    """
    generator = MaintenanceEstimateGenerator(
        kmz_path=kmz_path,
        buffer_distance_m=buffer_distance_m
    )

    generator.generate_estimate(
        tickets_df=tickets_df,
        output_path=output_path,
        project_name=project_name
    )


if __name__ == "__main__":
    # Test the maintenance estimate generator
    print("MaintenanceEstimateGenerator - Test Mode\n")

    kmz_path = Path("../projects/wink/route/wink.kmz")

    if kmz_path.exists():
        print(f"Loading route from {kmz_path}")
        generator = MaintenanceEstimateGenerator(kmz_path=kmz_path)

        print(f"\n✓ Loaded {len(generator.route_legs)} route legs:")
        for idx, leg in generator.route_legs.iterrows():
            print(f"  - {leg['name']}")
    else:
        print(f"Test file not found: {kmz_path}")
